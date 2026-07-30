from openpyxl import load_workbook
def register():
    print("WELCOME TO REGISTER PAGE")
    vef=input("If you are new user please register by entering register here or if you are already registered please enter login go to the Login page")
    if vef=="login":
        from login import login
        login()

    workbook=load_workbook("database/users.xlsx")
    sheet=workbook.active
    userid=sheet.max_row  

    username=input("Enter your username:")
    age=int(input("Enter your age:"))
    email=input("Enter your email:")
    address=input("Enter your address:")
    password=input("Enter your password:")

    sheet.append([userid,username,age,email,address,password])

    workbook.save("database/users.xlsx")

    print("Registration successful")

